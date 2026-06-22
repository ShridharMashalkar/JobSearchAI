"""
Author: Shridhar Mashalkar
This file is part of an open-source project licensed under the AGPL-3.0 License.
You are free to use, modify, and distribute this code under the terms of the LICENSE file.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from resume_tailor.config import AppConfig, ExcelSettings, FilenameSettings, LLMSettings, PDFSettings, PathSettings, PipelineSettings
from resume_tailor.excel import append_job_row, build_tracking_row
from resume_tailor.latex import render_cover_letter_tex, render_resume_tex
from resume_tailor.resume_data import load_resume_data
from resume_tailor.schemas import CoverLetterResult, ResumeTailoringResult, ScreeningResult
from resume_tailor.pipeline import run_pipeline
from resume_tailor.schemas import ResumeData, ResumeExperience, ResumeEducation, ResumeProject


def _load_config(base_dir: Path) -> AppConfig:
    return AppConfig(
        base_dir=base_dir,
        paths=PathSettings(
            jd_input_folder=base_dir / "jobs" / "input",
            output_folder=base_dir / "output",
            error_folder=base_dir / "jobs" / "error_jd",
            work_root=base_dir / ".work",
            resume_yaml_path=base_dir / "data" / "resume.yaml",
            resume_template_path=base_dir / "templateCV.tex",
            cover_letter_template_path=base_dir / "templateCoverLetter.tex",
            jobs_excel_path=base_dir / "tracking" / "jobs.xlsx",
        ),
        filenames=FilenameSettings(),
        llm=LLMSettings(),
        pdf=PDFSettings(),
        excel=ExcelSettings(),
        pipeline=PipelineSettings(),
    )


def _get_mock_resume_data() -> ResumeData:
    return ResumeData(
        name="John Doe",
        location="Munich, Germany",
        phone="+49 151 98765432",
        email="john.doe@example.com",
        linkedin_url="https://www.linkedin.com/in/johndoe-data/",
        linkedin_text="linkedin.com/in/johndoe-data",
        headline="Data Analytics Professional",
        summary="Results-driven Data Analytics professional with 4+ years of experience in technology and retail sectors. Proven expertise in Python, SQL, Power BI, data analysis, ETL, experimentation, and machine learning. Currently pursuing advanced studies in Business Analytics in Germany and authorized to work in the European Union.",
        skills={
            "technical": ["Data Analysis"],
            "tech_stack": ["Python"],
            "soft": ["Problem Solving"]
        },
        education=[
            ResumeEducation(
                date_range="09/2023 -- Present",
                degree="Master of Science (Business Analytics)",
                institution="Technical University of Munich",
                details="Merit Scholarship Recipient, Grade: 1.5 (German scale)",
                focus="Data analytics, machine learning, statistics"
            ),
            ResumeEducation(
                date_range="08/2015 -- 07/2019",
                degree="Bachelor of Engineering (Computer Engineering)",
                institution="Metropolitan Institute of Technology",
                details="Grade: 1.8 (German scale)",
                focus="Software Engineering, Databases, Networks"
            )
        ],
        experience=[
            ResumeExperience(
                date_range="05/2025 -- Present",
                location="Munich, Germany",
                title="Business Intelligence Working Student",
                company="TechNova Solutions GmbH",
                bullets=[
                    "Designed automated reporting pipelines connecting cloud databases and analytics platforms for enterprise reporting.",
                    "Generated operational insights that identified process inefficiencies and supported optimization initiatives across multiple teams.",
                    "Developed interactive dashboards for executive decision-making and resource planning."
                ],
                TotalBulletPointsToInclude=2
            ),
            ResumeExperience(
                date_range="09/2024 -- 02/2025",
                location="Munich, Germany",
                title="Research Assistant - Data Science",
                company="Technical University of Munich",
                bullets=[
                    "Mentored students in Python, machine learning, and statistical modeling concepts.",
                    "Conducted workshops covering data preprocessing, feature engineering, model evaluation, and visualization."
                ],
                TotalBulletPointsToInclude=1
            ),
            ResumeExperience(
                date_range="02/2021 -- 03/2023",
                location="Pune, India",
                title="Data Analyst",
                company="Insight Analytics Pvt. Ltd.",
                bullets=[
                    "Automated ETL workflows processing large-scale customer datasets from multiple business units.",
                    "Performed customer segmentation analysis to identify behavioral patterns and improve campaign effectiveness."
                ],
                TotalBulletPointsToInclude=4
            ),
            ResumeExperience(
                date_range="08/2020 -- 10/2020",
                location="Pune, India",
                title="Data Engineering Intern",
                company="DataSphere Technologies",
                bullets=[
                    "Transformed large datasets by parsing and cleaning structured and semi-structured data sources for analytics use cases.",
                    "Developed automated web data collection solutions using Python-based scraping frameworks."
                ],
                TotalBulletPointsToInclude=2
            )
        ],
        projects=[
            ResumeProject(
                name="Customer Churn Prediction Platform",
                description="Built a machine learning system to predict customer churn using publicly available datasets."
            ),
            ResumeProject(
                name="Retail Market Intelligence Dashboard",
                description="Developed a scalable data collection and visualization platform to analyze pricing trends and market dynamics."
            )
        ],
        certifications=[
            "Advanced Data Analytics Professional Certificate",
            "Cloud Data Engineering Fundamentals",
            "Machine Learning Specialization"
        ],
        publications=[
            "Data Governance and Process Automation in Enterprise Systems"
        ],
        languages=["Fluent English", "Intermediate German"],
        hobbies=[
            "I enjoy participating in technology meetups, hiking in nature, solving logic puzzles, and following international football."
        ]
    )




def test_renderers_keep_template_structure() -> None:
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
    )
    resume_tailoring = ResumeTailoringResult(
        professional_summary="Production AI engineer building reliable systems.",
        job_bullet_points={
            "TechNova Solutions": ["Tailored Bosch bullet"],
            "Research Assistant": ["Tailored student assistant bullet"],
        },
        technical_skills=["Python", "LangChain"],
        tech_stack=["FastAPI", "LangChain", "LangGraph", "Docker", "Azure"],
        soft_skills=["clear thinker"],
        masters_focus="Machine learning, deep learning, Python, data analytics",
        projects=[],
    )
    cover_letter = CoverLetterResult(
        subject="Subject: Application for AI Engineer — Acme AI",
        first_paragraph="First.",
        second_paragraph="Second.",
        third_paragraph="Third.",
        closing_paragraph="Closing.",
    )

    source_resume = (base_dir / "templateCV.tex").read_text(encoding="utf-8")
    rendered_resume = render_resume_tex(source_resume, resume_data, resume_tailoring, screening)
    assert "Production AI engineer building reliable systems." in rendered_resume
    assert "AI engineer with 3+ years of experience delivering Python, LLM, RAG, and LangChain solutions" not in rendered_resume
    assert "Tailored Bosch bullet" in rendered_resume
    assert "Tailored student assistant bullet" in rendered_resume
    assert "FastAPI, LangChain, LangGraph, Docker, Azure" in rendered_resume
    assert "Designed LLM-powered RAG pipelines in Python and LangChain" not in rendered_resume
    assert "Machine learning, deep learning, Python, data analytics" in rendered_resume
    assert "participating in technology meetups" in rendered_resume

    source_cover = (base_dir / "templateCoverLetter.tex").read_text(encoding="utf-8")
    rendered_cover = render_cover_letter_tex(source_cover, resume_data, screening, cover_letter)
    assert "Acme AI" in rendered_cover


def test_dynamic_experience_rendering() -> None:
    from resume_tailor.schemas import ResumeExperience
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
    )
    resume_tailoring = ResumeTailoringResult(
        professional_summary="Production AI engineer building reliable systems.",
        job_bullet_points={},
        technical_skills=[],
        tech_stack=[],
        soft_skills=[],
        masters_focus="",
        projects=[],
    )
    
    source_resume = (base_dir / "templateCV.tex").read_text(encoding="utf-8")

    # Case 1: 0 experiences
    resume_data_0 = resume_data.model_copy(update={"experience": []})
    rendered_resume = render_resume_tex(source_resume, resume_data_0, resume_tailoring, screening)
    assert rendered_resume.count("\\entry") == 2

    # Case 2: 1 experience
    exp = ResumeExperience(
        date_range="2025 -- 2026",
        location="Berlin",
        company="TechCorp",
        title="Software Dev",
        bullets=["Wrote code"],
    )
    resume_data_1 = resume_data.model_copy(update={"experience": [exp]})
    rendered_resume = render_resume_tex(source_resume, resume_data_1, resume_tailoring, screening)
    assert rendered_resume.count("\\entry") == 3
    assert "TechCorp" in rendered_resume
    assert "Berlin" in rendered_resume

    # Case 3: 5 experiences
    exps = [exp] * 5
    resume_data_5 = resume_data.model_copy(update={"experience": exps})
    rendered_resume = render_resume_tex(source_resume, resume_data_5, resume_tailoring, screening)
    assert rendered_resume.count("\\entry") == 7


def test_dynamic_education_rendering() -> None:
    from resume_tailor.schemas import ResumeEducation
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
    )
    resume_tailoring = ResumeTailoringResult(
        professional_summary="Production AI engineer building reliable systems.",
        job_bullet_points={},
        technical_skills=[],
        tech_stack=[],
        soft_skills=[],
        masters_focus="Tailored Focus",
        projects=[],
    )
    
    source_resume = (base_dir / "templateCV.tex").read_text(encoding="utf-8")

    # Case 1: 0 educations (should have 4 entry blocks since resume.yaml has 4 experiences by default)
    resume_data_0 = resume_data.model_copy(update={"education": []})
    rendered_resume = render_resume_tex(source_resume, resume_data_0, resume_tailoring, screening)
    assert rendered_resume.count("\\entry") == 4

    # Case 2: 1 education
    edu = ResumeEducation(
        date_range="2020 -- 2024",
        degree="B.Sc.",
        institution="Some Uni",
        details="Grade 1.0",
        focus="Math",
    )
    resume_data_1 = resume_data.model_copy(update={"education": [edu]})
    rendered_resume = render_resume_tex(source_resume, resume_data_1, resume_tailoring, screening)
    assert rendered_resume.count("\\entry") == 5
    assert "Some Uni" in rendered_resume
    assert "Tailored Focus" in rendered_resume

    # Case 3: 3 educations
    edus = [edu] * 3
    resume_data_3 = resume_data.model_copy(update={"education": edus})
    rendered_resume = render_resume_tex(source_resume, resume_data_3, resume_tailoring, screening)
    assert rendered_resume.count("\\entry") == 7


def test_optional_certifications_and_publications() -> None:
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
    )
    resume_tailoring = ResumeTailoringResult(
        professional_summary="Production AI engineer building reliable systems.",
        job_bullet_points={},
        technical_skills=[],
        tech_stack=[],
        soft_skills=[],
        masters_focus="",
        projects=[],
    )
    
    source_resume = (base_dir / "templateCV.tex").read_text(encoding="utf-8")

    # Case 1: Both present
    rendered_resume = render_resume_tex(source_resume, resume_data, resume_tailoring, screening)
    assert "CERTIFICATIONS" in rendered_resume
    assert "PUBLICATIONS" in rendered_resume
    assert "Advanced Data Analytics Professional Certificate" in rendered_resume
    assert "Data Governance and Process Automation in Enterprise Systems" in rendered_resume

    # Case 2: Certifications empty, Publications present
    resume_data_no_certs = resume_data.model_copy(update={"certifications": []})
    rendered_resume = render_resume_tex(source_resume, resume_data_no_certs, resume_tailoring, screening)
    assert "CERTIFICATIONS" not in rendered_resume
    assert "PUBLICATIONS" in rendered_resume

    # Case 3: Certifications present, Publications empty
    resume_data_no_pubs = resume_data.model_copy(update={"publications": []})
    rendered_resume = render_resume_tex(source_resume, resume_data_no_pubs, resume_tailoring, screening)
    assert "CERTIFICATIONS" in rendered_resume
    assert "PUBLICATIONS" not in rendered_resume

    # Case 4: Both empty
    resume_data_empty = resume_data.model_copy(update={"certifications": [], "publications": []})
    rendered_resume = render_resume_tex(source_resume, resume_data_empty, resume_tailoring, screening)
    assert "CERTIFICATIONS" not in rendered_resume
    assert "PUBLICATIONS" not in rendered_resume


def test_dynamic_cover_letter_subject_name() -> None:
    from resume_tailor.llm import generate_cover_letter
    from unittest.mock import MagicMock
    
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()
    resume_data.name = "John Doe"

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
    )
    resume_tailoring = ResumeTailoringResult(
        professional_summary="Production AI engineer building reliable systems.",
        job_bullet_points={},
        technical_skills=[],
        tech_stack=[],
        soft_skills=[],
        masters_focus="",
        projects=[],
    )
    config = _load_config(base_dir)

    from resume_tailor import llm
    original_run_json = llm._run_json_stage
    captured_system_prompt = None

    def fake_run_json_stage(model, system_prompt, human_prompt, schema):
        nonlocal captured_system_prompt
        captured_system_prompt = system_prompt
        return CoverLetterResult(
            subject="Test subject",
            first_paragraph="First.",
            second_paragraph="Second.",
            third_paragraph="Third.",
            closing_paragraph="Closing.",
        )

    llm._run_json_stage = fake_run_json_stage
    try:
        generate_cover_letter(MagicMock(), "jd text", screening, resume_tailoring, resume_data, config)
        assert captured_system_prompt is not None
        assert '[Job Title] – John Doe' in captured_system_prompt
    finally:
        llm._run_json_stage = original_run_json


def test_dynamic_input_context_in_prompt() -> None:
    from resume_tailor.llm import generate_resume_tailoring
    from unittest.mock import MagicMock
    
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()
    resume_data.name = "Test Candidate Name"
    resume_data.totalExperienceInYears = "3.5+"

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
    )
    config = _load_config(base_dir)

    from resume_tailor import llm
    original_run_json = llm._run_json_stage
    captured_system_prompt = None

    def fake_run_json_stage(model, system_prompt, human_prompt, schema):
        nonlocal captured_system_prompt
        captured_system_prompt = system_prompt
        return ResumeTailoringResult(
            professional_summary="Summary",
            job_bullet_points={},
            technical_skills=[],
            tech_stack=[],
            soft_skills=[],
            masters_focus="",
            projects=[],
        )

    llm._run_json_stage = fake_run_json_stage
    try:
        generate_resume_tailoring(MagicMock(), "jd text", screening, resume_data, config)
        assert captured_system_prompt is not None
        assert "INPUTCONTENTTOINSERT" not in captured_system_prompt
        assert "Name: Test Candidate Name" in captured_system_prompt
        assert "TechNova Solutions GmbH | Business Intelligence Working Student | 05/2025 -- Present" in captured_system_prompt
        assert "Experience (3.5+ Years)" in captured_system_prompt
    finally:
        llm._run_json_stage = original_run_json


def test_config_driven_section_inclusion() -> None:
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
    )
    resume_tailoring = ResumeTailoringResult(
        professional_summary="Production AI engineer building reliable systems.",
        job_bullet_points={},
        technical_skills=[],
        tech_stack=[],
        soft_skills=[],
        masters_focus="",
        projects=[],
    )
    
    source_resume = (base_dir / "templateCV.tex").read_text(encoding="utf-8")

    # Case 1: All flags True (default) -> sections should be present
    rendered = render_resume_tex(
        source_resume, resume_data, resume_tailoring, screening,
        include_projects=True,
        include_certifications=True,
        include_publications=True
    )
    assert "PROJECTS" in rendered
    assert "CERTIFICATIONS" in rendered
    assert "PUBLICATIONS" in rendered

    # Case 2: All flags False -> sections should be completely removed
    rendered = render_resume_tex(
        source_resume, resume_data, resume_tailoring, screening,
        include_projects=False,
        include_certifications=False,
        include_publications=False
    )
    assert "PROJECTS" not in rendered
    assert "CERTIFICATIONS" not in rendered
    assert "PUBLICATIONS" not in rendered


def test_excel_append_creates_workbook(tmp_path: Path) -> None:
    base_dir = Path(__file__).resolve().parents[1]
    config = _load_config(base_dir)
    config = config.model_copy(update={"paths": config.paths.model_copy(update={"jobs_excel_path": tmp_path / "jobs.xlsx"})})

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
    )
    row = build_tracking_row(source_jd=tmp_path / "jd.txt", output_folder=tmp_path / "out", screening=screening)
    append_job_row(config, row)

    workbook = load_workbook(config.paths.jobs_excel_path)
    sheet = workbook[config.excel.sheet_name]
    assert sheet.max_row == 2
    header = [cell.value for cell in sheet[1]]
    company_index = header.index("Company") + 1
    assert sheet.cell(row=2, column=company_index).value == "Acme AI"

    date_index = header.index("Date") + 1
    date_val = sheet.cell(row=2, column=date_index).value
    assert len(date_val) == 10
    assert date_val[2] == "-"
    assert date_val[5] == "-"


def test_excel_appends_blank_job_url_when_missing(tmp_path: Path) -> None:
    base_dir = Path(__file__).resolve().parents[1]
    config = _load_config(base_dir)
    config = config.model_copy(update={"paths": config.paths.model_copy(update={"jobs_excel_path": tmp_path / "jobs.xlsx"})})

    screening = ScreeningResult(
        company_name="Acme AI",
        job_title="AI Engineer",
        about_company="Acme AI builds AI systems.",
        role_summary="Build production AI systems.",
        key_responsibilities=["Ship models"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["MLOps"],
        overall_fit=8,
        job_url=None,
    )
    row = build_tracking_row(source_jd=tmp_path / "jd.txt", output_folder=tmp_path / "out", screening=screening)
    append_job_row(config, row)

    workbook = load_workbook(config.paths.jobs_excel_path)
    sheet = workbook[config.excel.sheet_name]
    header = [cell.value for cell in sheet[1]]
    assert "Job URL" in header
    job_url_index = header.index("Job URL") + 1
    assert sheet.cell(row=2, column=job_url_index).value in (None, "")


def test_cover_letter_renderer_strips_duplicate_salutation_and_signoff() -> None:
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()

    screening = ScreeningResult(
        company_name="DataNova Analytics GmbH",
        job_title="Data Analyst (Python & BI)",
        about_company="DataNova Analytics GmbH builds analytics products.",
        role_summary="Build data workflows.",
        key_responsibilities=["Analyze data"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["SQL"],
        overall_fit=7,
        job_url=None,
    )
    cover_letter = CoverLetterResult(
        subject="Application for Data Analyst (Python & BI) — DataNova Analytics GmbH",
        first_paragraph="Dear Hiring Team, I am applying for the role.",
        second_paragraph="Second paragraph.",
        third_paragraph="Third paragraph.",
        closing_paragraph="I would welcome the opportunity to discuss this further. Sincerely, John Doe",
    )

    template_text = (base_dir / "templateCoverLetter.tex").read_text(encoding="utf-8")
    rendered = render_cover_letter_tex(template_text, resume_data, screening, cover_letter)

    assert rendered.count("Dear Hiring Team,") == 1
    assert rendered.count("Sincerely,") == 1
    assert "Dear Hiring Team, I am applying for the role." not in rendered
    assert "Sincerely, John Doe" not in rendered


def test_cover_letter_renderer_collapses_repeated_closing_block() -> None:
    base_dir = Path(__file__).resolve().parents[1]
    resume_data = _get_mock_resume_data()

    screening = ScreeningResult(
        company_name="DataNova Analytics GmbH",
        job_title="Data Analyst (Python & BI)",
        about_company="DataNova Analytics GmbH builds analytics products.",
        role_summary="Build data workflows.",
        key_responsibilities=["Analyze data"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["SQL"],
        overall_fit=7,
        job_url=None,
    )
    closing_block = (
        "I would welcome the opportunity to discuss how my analytics background and Python-based workflow "
        "experience can support DataNova Analytics GmbH. Thank you for your time and consideration. "
        "I would welcome the opportunity to discuss how my analytics background and Python-based workflow "
        "experience can support DataNova Analytics GmbH. Thank you for your time and consideration."
    )
    cover_letter = CoverLetterResult(
        subject="Application for Data Analyst (Python & BI) — DataNova Analytics GmbH",
        first_paragraph="First paragraph.",
        second_paragraph="Second paragraph.",
        third_paragraph="Third paragraph.",
        closing_paragraph=closing_block,
    )

    template_text = (base_dir / "templateCoverLetter.tex").read_text(encoding="utf-8")
    rendered = render_cover_letter_tex(template_text, resume_data, screening, cover_letter)

    assert rendered.count("I would welcome the opportunity to discuss how my analytics background and Python-based workflow") == 1


def test_pipeline_saves_rendered_latex_files_in_output_folder(tmp_path: Path, monkeypatch) -> None:
    base_dir = Path(__file__).resolve().parents[1]
    input_dir = tmp_path / "jobs" / "input"
    output_dir = tmp_path / "output"
    error_dir = tmp_path / "jobs" / "error_jd"
    work_dir = tmp_path / ".work"
    tracking_dir = tmp_path / "tracking"
    input_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)
    error_dir.mkdir(parents=True)
    work_dir.mkdir(parents=True)
    tracking_dir.mkdir(parents=True)

    jd_path = input_dir / "sample.txt"
    jd_path.write_text(
        "Sample job description for testing rendered latex output files.",
        encoding="utf-8",
    )

    screening = ScreeningResult(
        company_name="Example Co",
        job_title="Example Role",
        about_company="Example Co builds example products.",
        role_summary="Example summary.",
        key_responsibilities=["Build systems"],
        tech_skills=["Python"],
        soft_skills=["communication"],
        missing_keywords=["SQL"],
        overall_fit=8,
    )
    resume_tailoring = ResumeTailoringResult(
        professional_summary="Example tailored summary.",
        job_bullet_points={},
        technical_skills=["Python"],
        tech_stack=["Python"],
        soft_skills=["communication"],
        masters_focus="Example focus",
        projects=[],
    )
    cover_letter = CoverLetterResult(
        subject="Example subject",
        first_paragraph="First paragraph.",
        second_paragraph="Second paragraph.",
        third_paragraph="Third paragraph.",
        closing_paragraph="Closing paragraph.",
    )

    def fake_call_stage_with_retries(*, stage_name, model, max_retries, start_retry, runner):
        if stage_name == "ATS screening":
            return screening
        if stage_name == "resume tailoring":
            return resume_tailoring
        if stage_name == "cover letter":
            return cover_letter
        raise AssertionError(stage_name)

    def fake_compile_latex(tex_path, output_pdf_path, config):
        output_pdf_path.write_text(f"PDF for {tex_path.name}", encoding="utf-8")
        return output_pdf_path

    config_path = tmp_path / "pipeline_config.yaml"
    config_path.write_text(
        "\n".join(
            [
                f"paths:",
                f"  jd_input_folder: {input_dir.as_posix()}",
                f"  output_folder: {output_dir.as_posix()}",
                f"  error_folder: {error_dir.as_posix()}",
                f"  work_root: {work_dir.as_posix()}",
                f"  resume_yaml_path: {(base_dir / 'data' / 'resume.yaml').as_posix()}",
                f"  resume_template_path: {(base_dir / 'templateCV.tex').as_posix()}",
                f"  cover_letter_template_path: {(base_dir / 'templateCoverLetter.tex').as_posix()}",
                f"  jobs_excel_path: {(tracking_dir / 'jobs.xlsx').as_posix()}",
                "filenames:",
                "  analysis_filename: AI_Analysis.txt",
                "  jd_filename: jd.txt",
                "  resume_tex_filename: resume.tex",
                "  cover_letter_tex_filename: cover_letter.tex",
                "  resume_pdf_filename: resume.pdf",
                "  cover_letter_pdf_filename: cover_letter.pdf",
                "  error_jd_suffix: _failed",
                "llm:",
                "  model_name: gpt-5.4-mini",
                "  temperature: 0.0",
                "  max_retries: 0",
                "  retry_start: 0",
                "  request_timeout_seconds: 60",
                "pdf:",
                "  command: pdflatex",
                "  runs: 1",
                "  timeout_seconds: 120",
                "  extra_args:",
                "    - -interaction=nonstopmode",
                "    - -halt-on-error",
                "    - -file-line-error",
                "excel:",
                "  sheet_name: Jobs",
                "  columns:",
                "    - Company",
                "    - Title",
                "    - Overall Fit",
                "    - Date",
                "    - Application Status",
                "    - Job URL",
                "    - Role Summary",
                "    - Key Responsibilities",
                "    - Technical Skills",
                "    - Soft Skills",
                "    - Missing Keywords",
                "    - Output Folder",
                "pipeline:",
                "  jd_glob: \"*.txt\"",
                "  job_folder_separator: \"_\"",
                "  folder_slug_max_length: 120",
                "  preserve_case: false",
                "  source_encoding: utf-8",
            ]
        ),
        encoding="utf-8",
    )

    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr("resume_tailor.pipeline.call_stage_with_retries", fake_call_stage_with_retries)
    monkeypatch.setattr("resume_tailor.pipeline.compile_latex", fake_compile_latex)
    run_pipeline(config_path)

    job_folders = list(output_dir.iterdir())
    assert job_folders, "expected a job output folder"
    job_folder = job_folders[0]
    assert (job_folder / "resume.tex").exists()
    assert (job_folder / "cover_letter.tex").exists()
    assert (job_folder / "developercv.cls").exists()
    assert (job_folder / "resume.pdf").exists()
    assert (job_folder / "cover_letter.pdf").exists()
