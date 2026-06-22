import os
import json
import logging
from http.server import HTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

#Command: python server.py

# Setup logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("dashboard_server")

# Try to find the Excel path from config
def get_excel_path():
    default_path = Path("tracking/jobs.xlsx")
    config_path = Path("config/pipeline_config.yaml")
    if config_path.exists():
        try:
            import yaml
            with open(config_path, "r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f)
                path_str = cfg.get("paths", {}).get("jobs_excel_path", "tracking/jobs.xlsx")
                return Path(path_str)
        except Exception:
            pass
    return default_path

def format_date_for_excel(val):
    if not val:
        return ""
    # JS input[type="date"] returns YYYY-MM-DD
    try:
        dt = datetime.strptime(val, "%Y-%m-%d")
        return dt.strftime("%d-%m-%Y")
    except ValueError:
        pass
    try:
        # If it's already in DD-MM-YYYY format
        dt = datetime.strptime(val, "%d-%m-%Y")
        return dt.strftime("%d-%m-%Y")
    except ValueError:
        pass
    return val

JS_TO_EXCEL_MAP = {
    "company": "Company",
    "title": "Title",
    "fit": "Overall Fit",
    "date": "Date",
    "status": "Application Status",
    "url": "Job URL",
    "summary": "Role Summary",
    "responsibilities": "Key Responsibilities",
    "techSkills": "Technical Skills",
    "softSkills": "Soft Skills",
    "missingKeywords": "Missing Keywords",
    "folder": "Output Folder",
    "customNotes": "Custom Notes",
}

class DashboardAPIHandler(SimpleHTTPRequestHandler):
    def end_headers(self):
        # Enable CORS and disable caching
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        self.send_header("X-Powered-By", "CareerTrack AI (Developed by Shridhar Mashalkar)")
        super().end_headers()

    def do_GET(self):
        # Intercept jobs.xlsx requests and serve it from the actual path
        path_without_query = self.path.split('?')[0]
        if path_without_query == "/jobs.xlsx":
            excel_path = get_excel_path()
            if excel_path.exists():
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Length", str(excel_path.stat().st_size))
                self.end_headers()
                with open(excel_path, "rb") as f:
                    self.wfile.write(f.read())
                return
            else:
                self.send_error(404, f"Excel file not found at {excel_path}")
                return
        
        super().do_GET()

    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()

    def do_POST(self):
        if self.path == "/api/update":
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            
            try:
                payload = json.loads(post_data.decode("utf-8"))
                excel_index = payload.get("excelIndex")
                
                if not excel_index or not isinstance(excel_index, int):
                    self._send_json({"success": False, "error": "Invalid excelIndex parameter"}, status=400)
                    return
                
                excel_path = get_excel_path()
                if not excel_path.exists():
                    self._send_json({"success": False, "error": f"Excel file not found at {excel_path.resolve()}"}, status=404)
                    return

                # Load workbook
                wb = load_workbook(excel_path)
                sheet_name = "Applications"
                excel_columns = [
                    "Company",
                    "Title",
                    "Overall Fit",
                    "Date",
                    "Application Status",
                    "Job URL",
                    "Role Summary",
                    "Key Responsibilities",
                    "Technical Skills",
                    "Soft Skills",
                    "Missing Keywords",
                    "Output Folder",
                    "Custom Notes",
                ]
                config_path = Path("config/pipeline_config.yaml")
                if config_path.exists():
                    try:
                        import yaml
                        with open(config_path, "r", encoding="utf-8") as f:
                            cfg = yaml.safe_load(f)
                            sheet_name = cfg.get("excel", {}).get("sheet_name", "Applications")
                            excel_columns = cfg.get("excel", {}).get("columns", excel_columns)
                    except Exception:
                        pass
                
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                else:
                    sheet = wb.active
                
                # Ensure all configured columns exist in the worksheet
                existing_headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
                if existing_headers and not all(header is None for header in existing_headers):
                    for column in excel_columns:
                        if column not in existing_headers:
                            sheet.cell(row=1, column=sheet.max_column + 1, value=column)
                            for r_idx in range(2, sheet.max_row + 1):
                                sheet.cell(row=r_idx, column=sheet.max_column, value="")
                            existing_headers.append(column)
                
                # Dynamic mapping based on spreadsheet headers
                headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
                header_to_col_idx = {h: idx + 1 for idx, h in enumerate(headers) if h is not None}
                
                for key, val in payload.items():
                    excel_header = JS_TO_EXCEL_MAP.get(key)
                    if excel_header and excel_header in header_to_col_idx:
                        col_idx = header_to_col_idx[excel_header]
                        
                        # Serialize list/dict values
                        if isinstance(val, (list, dict)):
                            cell_val = json.dumps(val, ensure_ascii=False)
                        elif key == "date" and val:
                            cell_val = format_date_for_excel(val)
                        elif key == "fit":
                            if val == "" or val is None:
                                cell_val = ""
                            else:
                                try:
                                    cell_val = int(val)
                                except ValueError:
                                    cell_val = val
                        else:
                            cell_val = val if val is not None else ""
                        
                        sheet.cell(row=excel_index, column=col_idx, value=cell_val)
                
                wb.save(excel_path)
                wb.close()
                
                logger.info(f"Successfully updated row {excel_index} in {excel_path}")
                self._send_json({"success": True})
            except Exception as e:
                logger.exception("Error processing update")
                self._send_json({"success": False, "error": str(e)}, status=500)
        else:
            super().do_POST()

    def _send_json(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.end_headers()
        self.wfile.write(json.dumps(data).encode("utf-8"))

def main():
    port = 8000
    server_address = ("", port)
    httpd = HTTPServer(server_address, DashboardAPIHandler)
    print("\n" + "=" * 60)
    print("      CareerTrack AI - Job Application Dashboard Server")
    print("           Developed by Shridhar Mashalkar")
    print("=" * 60 + "\n")
    logger.info(f"Starting dashboard local server on port {port}...")
    logger.info("Open http://localhost:8000/dashboard.html in your browser")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        logger.info("\nStopping server...")
        httpd.server_close()

if __name__ == "__main__":
    main()
