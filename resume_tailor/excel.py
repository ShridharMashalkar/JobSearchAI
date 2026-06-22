"""
Author: Shridhar Mashalkar
This file is part of an open-source project licensed under the AGPL-3.0 License.
You are free to use, modify, and distribute this code under the terms of the LICENSE file.
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from resume_tailor.config import AppConfig
from resume_tailor.files import ensure_directory
from resume_tailor.schemas import JobTrackingRow


def append_job_row(config: AppConfig, row: JobTrackingRow) -> None:
    workbook_path = config.paths.jobs_excel_path
    ensure_directory(workbook_path.parent)

    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
        sheet = workbook[config.excel.sheet_name] if config.excel.sheet_name in workbook.sheetnames else workbook.create_sheet(config.excel.sheet_name)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = config.excel.sheet_name
        _write_headers(sheet, config.excel.columns)

    _ensure_sheet_columns(sheet, config.excel.columns)

    sheet.append(_serialize_row(row, config.excel.columns))
    workbook.save(workbook_path)


def _write_headers(sheet: Worksheet, columns: list[str]) -> None:
    sheet.append(columns)


def _ensure_sheet_columns(sheet: Worksheet, columns: list[str]) -> None:
    existing_headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
    if not existing_headers or all(header is None for header in existing_headers):
        _write_headers(sheet, columns)
        return

    if existing_headers == columns:
        return

    missing_headers = [column for column in columns if column not in existing_headers]
    if not missing_headers:
        return

    for missing_header in missing_headers:
        sheet.cell(row=1, column=sheet.max_column + 1, value=missing_header)
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row=row_index, column=sheet.max_column, value="")


COLUMN_MAPPING = {
    "Company": "company_name",
    "Title": "job_title",
    "Overall Fit": "overall_fit",
    "Date": "processed_at",
    "Application Status": "status",
    "Job URL": "job_url",
    "Role Summary": "role_summary",
    "Key Responsibilities": "key_responsibilities",
    "Technical Skills": "tech_skills",
    "Soft Skills": "soft_skills",
    "Missing Keywords": "missing_keywords",
    "Output Folder": "output_folder",
    "Custom Notes": "custom_notes",
}


def _serialize_row(row: JobTrackingRow, columns: list[str]) -> list[object]:
    payload = row.model_dump()
    row_values = []
    for column in columns:
        field_name = COLUMN_MAPPING.get(column, column)
        value = payload.get(field_name, None)
        if (column == "Date" or field_name == "processed_at") and value:
            try:
                dt = datetime.fromisoformat(str(value))
                if dt.tzinfo is not None:
                    dt = dt.astimezone()
                value = dt.strftime("%d-%m-%Y")
            except ValueError:
                pass
        row_values.append(_serialize_value(value))
    return row_values


def _serialize_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def build_tracking_row(
    *,
    source_jd: Path,
    output_folder: Path,
    screening,
    processed_at: datetime | None = None,
    status: str = "CustomDocsGenerated",
    custom_notes: str = "",
) -> JobTrackingRow:
    moment = processed_at or datetime.now(timezone.utc)
    return JobTrackingRow(
        processed_at=moment.isoformat(),
        source_jd=str(source_jd),
        job_url=getattr(screening, "job_url", None),
        output_folder=str(output_folder),
        status=status,
        company_name=screening.company_name,
        job_title=screening.job_title,
        role_summary=screening.role_summary,
        key_responsibilities=screening.key_responsibilities,
        tech_skills=screening.tech_skills,
        soft_skills=screening.soft_skills,
        missing_keywords=screening.missing_keywords,
        overall_fit=screening.overall_fit,
        custom_notes=custom_notes,
    )