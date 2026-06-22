"""
Author: Shridhar Mashalkar
This file is part of an open-source project licensed under the AGPL-3.0 License.
You are free to use, modify, and distribute this code under the terms of the LICENSE file.
Terminal Command: python -m resume_tailor.cleanup
"""
from __future__ import annotations

import argparse
import logging
import shutil
from pathlib import Path

from openpyxl import load_workbook

from resume_tailor.config import load_config
from resume_tailor.files import ensure_directory, slugify, unique_path

LOGGER = logging.getLogger("resume_tailor.cleanup")


def cleanup_rejected_resumes(config_path: Path) -> None:
    config = load_config(config_path)
    excel_path = config.paths.jobs_excel_path

    if not excel_path.exists():
        LOGGER.warning("Excel file %s does not exist. No jobs to clean up.", excel_path)
        return

    rejected_dir = ensure_directory(config.base_dir / "rejected_resumes")

    workbook = load_workbook(excel_path)
    sheet_name = config.excel.sheet_name
    if sheet_name not in workbook.sheetnames:
        LOGGER.warning("Sheet %s not found in workbook.", sheet_name)
        return

    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []

    # Map column indexes
    status_idx = headers.index("Application Status") if "Application Status" in headers else -1
    folder_idx = headers.index("Output Folder") if "Output Folder" in headers else -1
    company_idx = headers.index("Company") if "Company" in headers else -1
    title_idx = headers.index("Title") if "Title" in headers else -1

    if status_idx == -1:
        LOGGER.warning("Column 'Application Status' not found in Excel headers.")
        return

    moved_count = 0

    for row in sheet.iter_rows(min_row=2):
        status_val = row[status_idx].value
        if not status_val:
            continue

        # Check if the status is Rejected or Cancelled
        status_str = str(status_val).strip().lower()
        if status_str in ("rejected", "cancelled"):
            company = row[company_idx].value if company_idx != -1 else "Unknown"
            title = row[title_idx].value if title_idx != -1 else "Untitled"

            # Find output folder candidates
            folder_candidates = []

            # 1. From the excel column
            if folder_idx != -1 and row[folder_idx].value:
                folder_candidates.append(Path(str(row[folder_idx].value)))

            # 2. Derive from slug
            if company_idx != -1 and title_idx != -1 and company and title:
                company_slug = slugify(
                    str(company),
                    separator=config.pipeline.job_folder_separator,
                    preserve_case=config.pipeline.preserve_case,
                    max_length=config.pipeline.folder_slug_max_length,
                )
                job_title_slug = slugify(
                    str(title),
                    separator=config.pipeline.job_folder_separator,
                    preserve_case=config.pipeline.preserve_case,
                    max_length=config.pipeline.folder_slug_max_length,
                )
                folder_name = f"{company_slug}{config.pipeline.job_folder_separator}{job_title_slug}"
                folder_candidates.append(config.paths.output_folder / folder_name)

            # Try to find an existing candidate
            source_dir = None
            for candidate in folder_candidates:
                # If path is relative, resolve it relative to base_dir
                if not candidate.is_absolute():
                    candidate = (config.base_dir / candidate).resolve()
                if candidate.exists() and candidate.is_dir():
                    source_dir = candidate
                    break

            if source_dir:
                dest_dir = unique_path(rejected_dir / source_dir.name)
                LOGGER.info("Moving %s job folder: %s -> %s", status_str, source_dir, dest_dir)
                try:
                    shutil.move(str(source_dir), str(dest_dir))
                    moved_count += 1
                except Exception as e:
                    LOGGER.error("Failed to move %s to %s: %s", source_dir, dest_dir, e)
            else:
                LOGGER.debug("No output folder found for %s job %s - %s", status_str, company, title)

    LOGGER.info("Cleanup completed. Moved %s rejected/cancelled job folders.", moved_count)


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")

    parser = argparse.ArgumentParser(description="Cleanup rejected/cancelled resumes and move their folders to rejected_resumes.")
    parser.add_argument("--config", type=Path, help="Path to the pipeline YAML config file.")
    args = parser.parse_args()

    config_path = args.config
    if not config_path:
        # Resolve default config path relative to the script location
        config_path = Path(__file__).parent.parent / "config" / "pipeline_config.yaml"

    if not config_path.exists():
        LOGGER.error("Config file not found at %s. Please specify --config.", config_path)
        return

    cleanup_rejected_resumes(config_path)


if __name__ == "__main__":
    main()
